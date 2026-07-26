from __future__ import annotations

import json
import mimetypes
from pathlib import Path

from backend.app.core.ai_config import (
    DEFAULT_COPILOT_MODEL,
    DEFAULT_EMBEDDING_MODEL,
    DEFAULT_OLLAMA_BASE_URL,
)
from backend.app.core.indexing_config import (
    DEFAULT_IGNORED_DIRECTORY_NAMES,
    DEFAULT_SUPPORTED_EXTENSIONS,
)
from backend.app.core.logging import logger

DEFAULT_SETTINGS: dict = {
    "ignoredDirectoryNames": sorted(DEFAULT_IGNORED_DIRECTORY_NAMES),
    "supportedExtensions": sorted(DEFAULT_SUPPORTED_EXTENSIONS),
    "autoIndexOnChange": True,
    "autoIndexOnStartup": True,
    "maxFileSizeMb": 50,
    "notifications": {
        "indexingComplete": True,
        "indexingErrors": True,
        "watcherEvents": False,
    },
    "embeddingProvider": "ollama",
    "embeddingModel": DEFAULT_EMBEDDING_MODEL,
    "ollamaBaseUrl": DEFAULT_OLLAMA_BASE_URL,
    "chunkSize": 800,
    "chunkOverlap": 120,
    "autoEmbedOnIndex": True,
    "copilotProvider": "ollama",
    "copilotModel": DEFAULT_COPILOT_MODEL,
}


def merge_settings(raw: dict | None) -> dict:
    base = json.loads(json.dumps(DEFAULT_SETTINGS))
    if not raw:
        return base
    for key, value in raw.items():
        if key == "notifications" and isinstance(value, dict):
            base["notifications"] = {**base["notifications"], **value}
        elif key in base:
            base[key] = value
    return base


def extract_file_metadata(path: str) -> dict:
    """Basic metadata extraction for indexed files (no AI / embeddings)."""
    file_path = Path(path)
    ext = file_path.suffix.lower()
    meta: dict = {"extension": ext, "fileName": file_path.name}

    mime, _ = mimetypes.guess_type(path)
    meta["mimeType"] = mime

    if ext in {".txt", ".md", ".markdown", ".json", ".csv", ".py", ".ts", ".tsx", ".js", ".jsx"}:
        try:
            text = file_path.read_text(encoding="utf-8", errors="ignore")
            meta["charCount"] = len(text)
            meta["lineCount"] = text.count("\n") + 1 if text else 0
            meta["preview"] = text[:2000]
        except OSError as exc:
            logger.debug("Text preview failed for %s: %s", path, exc)

    if ext == ".pdf":
        try:
            from pypdf import PdfReader

            reader = PdfReader(str(file_path))
            meta["pageCount"] = len(reader.pages)
        except Exception as exc:
            logger.debug("PDF metadata failed for %s: %s", path, exc)

    if ext == ".docx":
        try:
            from docx import Document

            doc = Document(str(file_path))
            meta["paragraphCount"] = len(doc.paragraphs)
        except Exception as exc:
            logger.debug("DOCX metadata failed for %s: %s", path, exc)

    if ext == ".xlsx":
        try:
            from openpyxl import load_workbook

            wb = load_workbook(str(file_path), read_only=True, data_only=True)
            meta["sheetCount"] = len(wb.sheetnames)
            meta["sheetNames"] = wb.sheetnames[:10]
            wb.close()
        except Exception as exc:
            logger.debug("XLSX metadata failed for %s: %s", path, exc)

    if ext in {".png", ".jpg", ".jpeg", ".gif", ".webp"}:
        meta["mediaType"] = "image"

    return meta
