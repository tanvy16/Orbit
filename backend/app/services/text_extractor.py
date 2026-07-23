from __future__ import annotations

import csv
import io
import json
from pathlib import Path

from backend.app.core.logging import logger


def extract_document_text(path: str) -> str:
    file_path = Path(path)
    if not file_path.is_file():
        return ""

    ext = file_path.suffix.lower()
    try:
        if ext in {".txt", ".md", ".markdown", ".csv", ".json", ".py", ".ts", ".tsx", ".js", ".jsx", ".java", ".go", ".rs", ".cpp", ".c", ".h", ".css", ".scss", ".yaml", ".yml", ".log"}:
            return file_path.read_text(encoding="utf-8", errors="ignore")

        if ext == ".html":
            from bs4 import BeautifulSoup

            html = file_path.read_text(encoding="utf-8", errors="ignore")
            return BeautifulSoup(html, "html.parser").get_text(separator="\n")

        if ext == ".xml":
            import xml.etree.ElementTree as ET

            tree = ET.parse(file_path)
            return ET.tostring(tree.getroot(), encoding="unicode", method="text")

        if ext == ".json":
            data = json.loads(file_path.read_text(encoding="utf-8", errors="ignore"))
            return json.dumps(data, indent=2)

        if ext == ".csv":
            raw = file_path.read_text(encoding="utf-8", errors="ignore")
            reader = csv.reader(io.StringIO(raw))
            return "\n".join(", ".join(row) for row in reader)

        if ext == ".pdf":
            from pypdf import PdfReader

            reader = PdfReader(str(file_path))
            return "\n".join(page.extract_text() or "" for page in reader.pages)

        if ext == ".docx":
            from docx import Document

            doc = Document(str(file_path))
            return "\n".join(p.text for p in doc.paragraphs if p.text)

        if ext == ".xlsx":
            from openpyxl import load_workbook

            wb = load_workbook(str(file_path), read_only=True, data_only=True)
            lines: list[str] = []
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                lines.append(f"# Sheet: {sheet}")
                for row in ws.iter_rows(values_only=True):
                    lines.append("\t".join("" if c is None else str(c) for c in row))
            wb.close()
            return "\n".join(lines)

        if ext in {".png", ".jpg", ".jpeg", ".gif", ".webp"}:
            return ""

    except Exception as exc:
        logger.warning("Text extraction failed for %s: %s", path, exc)
        return ""

    return file_path.read_text(encoding="utf-8", errors="ignore") if ext else ""
